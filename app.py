from openpyxl import load_workbook

wb = load_workbook('events.xlsx')
ws = wb.active
if 'duplicate_ids' in wb.sheetnames:
    print('duplicate_ids exists')
    ws2 = wb['duplicate_ids']
else:
   ws2 = wb.create_sheet(title='duplicate_ids')


if 'missing_ids' in wb.sheetnames:
    print('missing_ids exists')
    ws3 = wb['missing_ids']
else:
   ws3 = wb.create_sheet(title='missing_ids')


data = []

print(ws)

for row in ws.iter_rows(min_row=2,min_col=9,max_col=9,max_row=1080,values_only=True):
    r = row[0]
    data.append(int(r[-4:]))

print(data)


xs = data
s = set()
any(x in s or s.add(x) for x in xs)
# You can use a similar approach to actually retrieve the duplicates.
s = set()
duplicates = set(x for x in xs if x in s or s.add(x))
print('found duplicates')
duplicates = sorted(list(duplicates))


    
ws2.cell(row=1,column=1,value='DUPLICATE IDS')
for r in duplicates:
    i = duplicates.index(r)
    print(f'{r}-{i}')
    ws2.cell(row=(i+2),column=1,value=f'{r}')


print(duplicates)
print(len(duplicates))


def find_missing(lst):
    return [i for x, y in zip(lst, lst[1:]) 
        for i in range(x + 1, y) if y - x > 1]
  
# Driver code
missing_ids = data
missing_ids = find_missing(missing_ids)
print(missing_ids)
print(len(missing_ids))

ws3.cell(row=1,column=1,value='MISSING IDS')
for rM in missing_ids:
    iM = missing_ids.index(rM)
    print(f'{rM}-{iM}')
    ws3.cell(row=(iM+2),column=1,value=f'{rM}')

wb.save('events.xlsx')
